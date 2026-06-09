import os
import re
import tempfile
import threading
import tkinter as tk
from tkinter import filedialog, messagebox
from concurrent.futures import ThreadPoolExecutor, as_completed

import qrcode
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from barcode import Code128
from barcode.writer import ImageWriter


# =========================
# 并行参数
# =========================

CPU_COUNT = os.cpu_count() or 4
MAX_WORKERS = min(8, CPU_COUNT)


# =========================
# 图片尺寸
# =========================

BARCODE_WIDTH = 320
BARCODE_HEIGHT = 90

QRCODE_DISPLAY_SIZE = 150
QRCODE_REAL_SIZE = 360


# =========================
# Excel 列宽
# =========================

COLUMN_A_WIDTH = 20
COLUMN_B_WIDTH = 45
COLUMN_C_WIDTH = 24


try:
    RESAMPLE_NEAREST = PILImage.Resampling.NEAREST
except AttributeError:
    RESAMPLE_NEAREST = PILImage.NEAREST


def remove_existing_images_in_bc(ws):
    """
    删除 B、C 列已有图片，避免重复生成时图片叠加。
    openpyxl 内部列号从 0 开始：A=0, B=1, C=2
    """
    keep_images = []

    for img in ws._images:
        col = None

        if isinstance(img.anchor, str):
            match = re.match(r"([A-Z]+)", img.anchor)
            if match:
                col_letter = match.group(1).upper()
                col = ord(col_letter) - ord("A")

        elif hasattr(img.anchor, "_from"):
            col = img.anchor._from.col

        if col not in (1, 2):
            keep_images.append(img)

    ws._images = keep_images


def prepare_header(ws):
    """
    增加表头：
    A1 = 编码
    B1 = 条形码
    C1 = 二维码
    """
    a1 = str(ws.cell(row=1, column=1).value or "").strip()
    b1 = str(ws.cell(row=1, column=2).value or "").strip()
    c1 = str(ws.cell(row=1, column=3).value or "").strip()

    has_header = (
        a1 in ["编码", "编号", "条码", "字符串", "数据", "code", "Code", "CODE"]
        or b1 in ["条形码", "barcode", "Barcode", "BARCODE"]
        or c1 in ["二维码", "QRCode", "QR Code", "qrcode"]
    )

    if not has_header:
        ws.insert_rows(1)

    ws.cell(row=1, column=1).value = "编码"
    ws.cell(row=1, column=2).value = "条形码"
    ws.cell(row=1, column=3).value = "二维码"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="999999")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, 4):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def make_barcode_png(text: str, output_path: str):
    """
    生成 Code128 条形码 PNG。
    注意：Code128 不支持中文。
    """
    code = Code128(text, writer=ImageWriter())

    options = {
        "module_width": 0.28,
        "module_height": 22,
        "quiet_zone": 3,
        "write_text": False,
    }

    code.save(output_path[:-4], options=options)


def make_qrcode_png(text: str, output_path: str):
    """
    生成高清二维码 PNG。
    二维码支持中文。
    """
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=12,
        border=4,
    )

    qr.add_data(text)
    qr.make(fit=True)

    img = qr.make_image(
        fill_color="black",
        back_color="white"
    ).convert("RGB")

    img = img.resize(
        (QRCODE_REAL_SIZE, QRCODE_REAL_SIZE),
        RESAMPLE_NEAREST
    )

    img.save(output_path)


def generate_images_for_row(row, text, barcode_path, qrcode_path):
    """
    单行图片生成任务。
    并行执行。
    条形码失败不影响二维码生成。
    """
    barcode_error = None
    qrcode_error = None

    try:
        make_barcode_png(text, barcode_path)
    except Exception as e:
        barcode_error = str(e)

    try:
        make_qrcode_png(text, qrcode_path)
    except Exception as e:
        qrcode_error = str(e)

    return {
        "row": row,
        "barcode_path": barcode_path,
        "qrcode_path": qrcode_path,
        "barcode_error": barcode_error,
        "qrcode_error": qrcode_error,
    }


def process_excel(input_path: str, output_path: str, progress_callback=None):
    """处理 Excel 文件"""
    wb = load_workbook(input_path)
    ws = wb.active

    remove_existing_images_in_bc(ws)
    prepare_header(ws)

    ws.column_dimensions["A"].width = COLUMN_A_WIDTH
    ws.column_dimensions["B"].width = COLUMN_B_WIDTH
    ws.column_dimensions["C"].width = COLUMN_C_WIDTH

    align_center = Alignment(horizontal="center", vertical="center")
    align_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="DDDDDD")
    normal_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row_items = []

    with tempfile.TemporaryDirectory() as temp_dir:
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row=row, column=1).value

            if value is None:
                continue

            text = str(value).strip()

            if not text:
                continue

            # 文件名只用行号，避免中文或特殊字符导致路径问题
            barcode_path = os.path.join(temp_dir, f"row_{row}_barcode.png")
            qrcode_path = os.path.join(temp_dir, f"row_{row}_qrcode.png")

            row_items.append({
                "row": row,
                "text": text,
                "barcode_path": barcode_path,
                "qrcode_path": qrcode_path,
            })

        total = len(row_items)

        if total == 0:
            wb.save(output_path)
            return {
                "total": 0,
                "barcode_failed": 0,
                "qrcode_failed": 0,
            }

        image_results = {}

        workers = min(MAX_WORKERS, total)

        if progress_callback:
            progress_callback(f"正在并行生成图片：0/{total}")

        with ThreadPoolExecutor(max_workers=workers) as executor:
            future_map = {}

            for item in row_items:
                future = executor.submit(
                    generate_images_for_row,
                    item["row"],
                    item["text"],
                    item["barcode_path"],
                    item["qrcode_path"]
                )
                future_map[future] = item["row"]

            done_count = 0

            for future in as_completed(future_map):
                result = future.result()
                image_results[result["row"]] = result

                done_count += 1

                if progress_callback:
                    progress_callback(f"正在并行生成图片：{done_count}/{total}")

        barcode_failed = 0
        qrcode_failed = 0

        if progress_callback:
            progress_callback("正在写入 Excel...")

        for index, item in enumerate(row_items, start=1):
            row = item["row"]
            result = image_results[row]

            ws.cell(row=row, column=1).alignment = align_center
            ws.cell(row=row, column=1).border = normal_border

            ws.cell(row=row, column=2).value = None
            ws.cell(row=row, column=3).value = None

            ws.cell(row=row, column=2).alignment = align_wrap
            ws.cell(row=row, column=3).alignment = align_center

            ws.cell(row=row, column=2).border = normal_border
            ws.cell(row=row, column=3).border = normal_border

            # 条形码：如果生成成功则插入图片，否则写提示文字
            if result["barcode_error"] is None and os.path.exists(result["barcode_path"]):
                barcode_img = ExcelImage(result["barcode_path"])
                barcode_img.width = BARCODE_WIDTH
                barcode_img.height = BARCODE_HEIGHT
                ws.add_image(barcode_img, f"B{row}")
            else:
                barcode_failed += 1
                ws.cell(row=row, column=2).value = "无法生成条形码\n含中文/非Code128字符"

            # 二维码：中文也可以生成
            if result["qrcode_error"] is None and os.path.exists(result["qrcode_path"]):
                qrcode_img = ExcelImage(result["qrcode_path"])
                qrcode_img.width = QRCODE_DISPLAY_SIZE
                qrcode_img.height = QRCODE_DISPLAY_SIZE
                ws.add_image(qrcode_img, f"C{row}")
            else:
                qrcode_failed += 1
                ws.cell(row=row, column=3).value = "二维码生成失败"

            image_height = max(BARCODE_HEIGHT, QRCODE_DISPLAY_SIZE)
            ws.row_dimensions[row].height = image_height * 0.75 + 20

            if progress_callback:
                progress_callback(f"正在写入 Excel：{index}/{total}")

        ws.auto_filter.ref = f"A1:C{ws.max_row}"

        if progress_callback:
            progress_callback("正在保存文件...")

        wb.save(output_path)

    return {
        "total": total,
        "barcode_failed": barcode_failed,
        "qrcode_failed": qrcode_failed,
    }


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel 条形码 / 二维码生成工具")
        self.root.geometry("700x270")
        self.root.resizable(False, False)

        self.input_path_var = tk.StringVar()
        self.output_path_var = tk.StringVar()
        self.status_var = tk.StringVar(value="请选择 Excel 文件")

        self.build_ui()

    def build_ui(self):
        tk.Label(self.root, text="Excel 文件：").place(x=20, y=25)

        tk.Entry(
            self.root,
            textvariable=self.input_path_var,
            width=70
        ).place(x=100, y=25)

        tk.Button(
            self.root,
            text="选择文件",
            command=self.select_input_file
        ).place(x=610, y=21)

        tk.Label(self.root, text="输出文件：").place(x=20, y=75)

        tk.Entry(
            self.root,
            textvariable=self.output_path_var,
            width=70
        ).place(x=100, y=75)

        tk.Button(
            self.root,
            text="另存为",
            command=self.select_output_file
        ).place(x=610, y=71)

        self.generate_button = tk.Button(
            self.root,
            text="开始生成",
            width=22,
            height=2,
            command=self.start_generate
        )
        self.generate_button.place(x=260, y=130)

        tk.Label(
            self.root,
            textvariable=self.status_var,
            fg="blue",
            wraplength=650,
            justify="left"
        ).place(x=20, y=210)

    def select_input_file(self):
        path = filedialog.askopenfilename(
            title="请选择 Excel 文件",
            filetypes=[
                ("Excel 文件", "*.xlsx"),
                ("所有文件", "*.*")
            ]
        )

        if path:
            self.input_path_var.set(path)

            folder = os.path.dirname(path)
            filename = os.path.basename(path)
            name, _ = os.path.splitext(filename)

            output_path = os.path.join(folder, f"{name}_output.xlsx")
            self.output_path_var.set(output_path)

            self.status_var.set("已选择文件，点击开始生成")

    def select_output_file(self):
        path = filedialog.asksaveasfilename(
            title="保存输出文件",
            defaultextension=".xlsx",
            filetypes=[
                ("Excel 文件", "*.xlsx")
            ]
        )

        if path:
            self.output_path_var.set(path)

    def start_generate(self):
        input_path = self.input_path_var.get().strip()
        output_path = self.output_path_var.get().strip()

        if not input_path:
            messagebox.showwarning("提示", "请先选择 Excel 文件")
            return

        if not os.path.exists(input_path):
            messagebox.showerror("错误", "输入文件不存在")
            return

        if not input_path.lower().endswith(".xlsx"):
            messagebox.showerror("错误", "目前仅支持 .xlsx 文件")
            return

        if not output_path:
            folder = os.path.dirname(input_path)
            filename = os.path.basename(input_path)
            name, _ = os.path.splitext(filename)
            output_path = os.path.join(folder, f"{name}_output.xlsx")
            self.output_path_var.set(output_path)

        if not output_path.lower().endswith(".xlsx"):
            output_path += ".xlsx"
            self.output_path_var.set(output_path)

        self.generate_button.config(state="disabled")
        self.status_var.set("正在处理，请稍候...")

        thread = threading.Thread(
            target=self.generate_worker,
            args=(input_path, output_path),
            daemon=True
        )
        thread.start()

    def update_status_safe(self, message):
        self.root.after(
            0,
            lambda msg=message: self.status_var.set(msg)
        )

    def generate_worker(self, input_path, output_path):
        try:
            result = process_excel(
                input_path,
                output_path,
                progress_callback=self.update_status_safe
            )

            self.root.after(
                0,
                lambda path=output_path, res=result: self.generate_success(path, res)
            )

        except Exception as e:
            error_msg = str(e)

            self.root.after(
                0,
                lambda msg=error_msg: self.generate_failed(msg)
            )

    def generate_success(self, output_path, result):
        self.generate_button.config(state="normal")

        total = result.get("total", 0)
        barcode_failed = result.get("barcode_failed", 0)
        qrcode_failed = result.get("qrcode_failed", 0)

        msg = (
            f"生成完成：{output_path}\n"
            f"总处理数量：{total}\n"
            f"条形码失败数量：{barcode_failed}\n"
            f"二维码失败数量：{qrcode_failed}"
        )

        self.status_var.set(msg)
        messagebox.showinfo("完成", msg)

    def generate_failed(self, error_msg):
        self.generate_button.config(state="normal")
        self.status_var.set("生成失败")
        messagebox.showerror("错误", f"生成失败：\n{error_msg}")


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
